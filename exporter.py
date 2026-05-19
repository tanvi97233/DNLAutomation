"""
Excel export with formatting.

Writes a DNL Report sheet (frozen, filtered, hyperlinked, HOT-highlighted)
plus a Summary sheet with counts.
"""
import os
from datetime import date, datetime
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "output"

# Column order required by spec
COLUMNS = [
    ("Serial Number", "serial_number", 8),
    ("Company Name", "company", 22),
    ("Date", "date", 13),
    ("News Type", "news_type", 18),
    ("Headline", "headline", 60),
    ("Source Link", "url", 40),
    ("Source Type", "source_type", 14),
    ("Hot vs Non-Hot", "hot", 14),
    ("Date Collected", "date_collected", 14),
]

NAVY = "1B3A5C"
HOT_YELLOW = "FFF2CC"
ALT_GREY = "F5F7FA"

_THIN = Side(style="thin", color="D0D7DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def _write_header(ws):
    """Style and write the header row."""
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=NAVY)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, record: Dict, is_hot: bool, alt: bool):
    """Write a single data row with styling."""
    align = Alignment(vertical="center", wrap_text=True)
    fill = None
    if is_hot:
        fill = PatternFill("solid", fgColor=HOT_YELLOW)
    elif alt:
        fill = PatternFill("solid", fgColor=ALT_GREY)

    for col_idx, (_label, key, _w) in enumerate(COLUMNS, start=1):
        value = record.get(key, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = align
        cell.border = _BORDER
        if fill:
            cell.fill = fill
        # Hyperlink styling for Source Link
        if key == "url" and value:
            cell.hyperlink = value
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            cell.font = Font(name="Calibri", size=10, bold=is_hot and key == "hot")

    ws.row_dimensions[row_idx].height = 45


def _write_summary(ws, records: List[Dict], start_date, end_date):
    """Populate the Summary sheet."""
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=NAVY)
    label_font = Font(name="Calibri", size=11, bold=True)

    ws["A1"] = "DNL Report — Summary"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Date Range"
    ws["B3"] = f"{start_date} to {end_date}"
    ws["A4"] = "Generated"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "Total Records"
    ws["B5"] = len(records)
    ws["A6"] = "HOT Items"
    ws["B6"] = sum(1 for r in records if r.get("hot") == "HOT")

    for row in range(3, 7):
        ws.cell(row=row, column=1).font = label_font

    ws["A8"] = "News Type"
    ws["B8"] = "Count"
    ws["A8"].font = label_font
    ws["B8"].font = label_font
    ws["A8"].fill = PatternFill("solid", fgColor="E8EEF7")
    ws["B8"].fill = PatternFill("solid", fgColor="E8EEF7")

    counts: Dict[str, int] = {}
    for r in records:
        nt = r.get("news_type") or "Other"
        counts[nt] = counts.get(nt, 0) + 1

    row_idx = 9
    for nt, c in sorted(counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row_idx, column=1, value=nt)
        ws.cell(row=row_idx, column=2, value=c)
        row_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16


def export_to_excel(records: List[Dict], start_date: date, end_date: date) -> str:
    """Write the formatted workbook and return its absolute path."""
    _ensure_output_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"DNL_{start_date}_{end_date}_{ts}.xlsx"
    filepath = os.path.abspath(os.path.join(OUTPUT_DIR, filename))

    wb = Workbook()
    ws = wb.active
    ws.title = "DNL Report"

    _write_header(ws)

    for i, rec in enumerate(records):
        is_hot = rec.get("hot") == "HOT"
        _write_row(ws, row_idx=i + 2, record=rec, is_hot=is_hot, alt=(i % 2 == 1))

    # Auto-filter
    last_col = get_column_letter(len(COLUMNS))
    last_row = max(1, len(records) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Summary sheet
    summary = wb.create_sheet("Summary")
    _write_summary(summary, records, start_date, end_date)

    wb.save(filepath)
    return filepath
